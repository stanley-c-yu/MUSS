import numpy as np
import pandas as pd
import os
import matplotlib.pyplot as plt
from matplotlib import rcParams, rcParamsDefault
import matplotlib.gridspec as gridspec
import seaborn as sns
from sklearn.metrics import f1_score
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from helper.annotation_processor import get_pa_labels, get_pa_abbr_labels
from helper.utils import generate_run_folder, strip_path
from clize import run
import logging


def format_for_excel(df, highlight_header=True):
    df = df.round(3)
    wb = Workbook()
    ws = wb.active
    font_style = Font(name='Times New Roman', size=12)
    # redFill = PatternFill(start_color='EE1111',
    #                       end_color='EE1111',
    #                       fill_type='solid')
    # greenFill = PatternFill(start_color='11EE11',
    #                         end_color='11EE11',
    #                         fill_type='solid')
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # whole sheet style
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font_style
            if 'nan' in str(cell.value):
                cell.value = cell.value.split(' ± nan')[0]
            # first row
            if cell.row == 1:
                cell.font = cell.font + Font(bold=True)
                cell.border = Border(
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if cell.column == 'A':
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            # first column
            elif cell.column == 'A':
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )

    # for col in range(2, ws.max_column + 1):
    #     col_letter = get_column_letter(col)
    #     ws.conditional_formatting.add(col_letter + '2:' + col_letter + str(ws.max_row), FormulaRule(
    #         formula=[col_letter + '2>=LARGE($' + col_letter + '$2:$' + col_letter + '$' + str(ws.max_row) + ',5)'], fill=greenFill))
    #     ws.conditional_formatting.add(col_letter + '2:' + col_letter + str(ws.max_row), FormulaRule(
    #         formula=[col_letter + '2<=SMALL($' + col_letter + '$2:$' + col_letter + '$' + str(ws.max_row) + ',5)'], fill=redFill))

    return wb


def top_and_bottom_n(df, column, n=5):
    top_n = df.nlargest(n, column)
    bottom_n = df.nsmallest(n, column)
    return pd.concat((top_n, bottom_n))

def top_n_misclassified_classes(conf_df, label, n=3):
    label_counts = conf_df.loc[label,:]
    label_percent = round(label_counts / sum(label_counts) * 100, 1)
    label_percent = label_percent.astype(str) + '%'
    label_df = pd.concat((label_percent, label_counts), axis=1)
    label_df.columns = ['%% of samples', '# of samples']
    label_df = label_df.loc[label_df.index != label,:]
    label_df = label_df.nlargest(n=3, columns=['# of samples']).reset_index(drop=False)
    return label_df.astype(str).apply(lambda row: row[0] + ', ' + row[1] + ' (' + row[2] + ')', axis=1)

def basic_stat(df, columns, method='min_max'):
    if method == 'min_max':
        result = df[columns].min().round(2).astype(str) + ' - ' + df[columns].max().round(2).astype(str)
        sort_col = df[columns[0]].mean().round(2)
        result['SORT'] = sort_col
        return result
    elif method == 'mean_std':
        result = df[columns].mean().round(2).astype(str) + ' ± ' + df[columns].std().round(2).astype(str)
        sort_col = df[columns[0]].mean().round(2)
        result['SORT'] = sort_col
        return result

def table_3(input_folder, debug=False):
    output_folder, metrics_file, _, _ = prepare_paths(input_folder, debug=debug)
    output_filepath = os.path.join(output_folder,'figures_and_tables', 'table3.csv')
    output_filepath_excel = os.path.join(output_folder,'figures_and_tables', 'table3.xlsx')
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    summary = pd.read_csv(metrics_file)
    filter_condition = (summary['FEATURE_TYPE'] == 'MO') & (
        summary['NUM_OF_SENSORS'] <= 3)
    table3_data = summary.loc[filter_condition, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'POSTURE_AVERAGE', 'LYING_POSTURE', 'SITTING_POSTURE', 'UPRIGHT_POSTURE']]

    table3_data = table3_data.sort_values(by=['POSTURE_AVERAGE'], ascending=False).drop_duplicates()
    # best models
    best_models = table3_data.groupby('NUM_OF_SENSORS').apply(lambda rows: rows.head(1)).reset_index(drop=True).sort_values(by=['POSTURE_AVERAGE'], ascending=False)
    best_models = best_models.loc[:, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'POSTURE_AVERAGE', 'LYING_POSTURE', 'SITTING_POSTURE', 'UPRIGHT_POSTURE']].round(2)
    
    # best models using wrist sensors
    table3_data['USE_DW'] = table3_data['SENSOR_PLACEMENT'].transform(
        lambda s: 'DW' in s.split('_'))
    table3_data['USE_NDW'] = table3_data['SENSOR_PLACEMENT'].transform(
        lambda s: 'NDW' in s.split('_'))
    table3_data['USE_W'] = table3_data['USE_DW'] | table3_data['USE_NDW']
    best_wrist_models = table3_data.loc[table3_data['USE_W'] == True,:].groupby('NUM_OF_SENSORS').apply(lambda rows: rows.nlargest(1, columns='POSTURE_AVERAGE')).reset_index(drop=True).sort_values(by=['POSTURE_AVERAGE'], ascending=False)
    best_wrist_models = best_wrist_models.loc[:, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'POSTURE_AVERAGE', 'LYING_POSTURE', 'SITTING_POSTURE', 'UPRIGHT_POSTURE']].round(2)
    # categorized performances
    table3_data['CATEGORY'] = 'W'
    
    # condition 1: A, H, T
    c1 = table3_data['SENSOR_PLACEMENT'].str.contains('A') & table3_data['SENSOR_PLACEMENT'].str.contains('H') & table3_data['SENSOR_PLACEMENT'].str.contains('T')
    table3_data.loc[c1, 'CATEGORY'] = 'A, H, T'
    # condition 2: A, T
    c2 = table3_data['SENSOR_PLACEMENT'].str.contains('A') & table3_data['SENSOR_PLACEMENT'].str.contains('T') & (~table3_data['SENSOR_PLACEMENT'].str.contains('H'))
    table3_data.loc[c2, 'CATEGORY'] = 'A, T'
    # condition 3: A, H
    c3 = table3_data['SENSOR_PLACEMENT'].str.contains('A') & table3_data['SENSOR_PLACEMENT'].str.contains('H') & (~table3_data['SENSOR_PLACEMENT'].str.contains('T'))
    table3_data.loc[c3, 'CATEGORY'] = 'A, H'

    # condition 4: H, T
    c4 = table3_data['SENSOR_PLACEMENT'].str.contains('T') & table3_data['SENSOR_PLACEMENT'].str.contains('H') & (~table3_data['SENSOR_PLACEMENT'].str.contains('A'))
    table3_data.loc[c4, 'CATEGORY'] = 'H, T'

    # condition 5: A
    c5 = table3_data['SENSOR_PLACEMENT'].str.contains('A') & (~table3_data['SENSOR_PLACEMENT'].str.contains('H')) & (~table3_data['SENSOR_PLACEMENT'].str.contains('T'))
    table3_data.loc[c5, 'CATEGORY'] = 'A'

    # condition 6: H
    c6 = table3_data['SENSOR_PLACEMENT'].str.contains('H') & (~table3_data['SENSOR_PLACEMENT'].str.contains('T')) & (~table3_data['SENSOR_PLACEMENT'].str.contains('A'))
    table3_data.loc[c6, 'CATEGORY'] = 'H'

    # condition 7: T
    c7 = table3_data['SENSOR_PLACEMENT'].str.contains('T') & (~table3_data['SENSOR_PLACEMENT'].str.contains('H')) & (~table3_data['SENSOR_PLACEMENT'].str.contains('A'))
    table3_data.loc[c7, 'CATEGORY'] = 'T'

    # condition 8: only W
    c8 = table3_data['SENSOR_PLACEMENT'].str.contains('W') & (~table3_data['SENSOR_PLACEMENT'].str.contains('H')) & (~table3_data['SENSOR_PLACEMENT'].str.contains('A')) & (~table3_data['SENSOR_PLACEMENT'].str.contains('T'))
    table3_data.loc[c8, 'CATEGORY'] = 'W only'
    
    result = table3_data.groupby(['NUM_OF_SENSORS', 'CATEGORY', 'USE_W']).apply(basic_stat, columns=['POSTURE_AVERAGE', 'LYING_POSTURE', 'SITTING_POSTURE', 'UPRIGHT_POSTURE'], method='mean_std').reset_index(drop=False).sort_values(['SORT', 'CATEGORY', 'NUM_OF_SENSORS'], ascending=False).drop(columns=['SORT'])
    
    result['CATEGORY'] = result['CATEGORY'] + result['USE_W'].transform(lambda x: ' (W)' if x else '')
    result = result.drop(columns=['USE_W'])

    best_models.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Lying', 'Sitting', 'Upright']
    best_wrist_models.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Lying', 'Sitting', 'Upright']
    result.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Lying', 'Sitting', 'Upright']

    best_models.loc[:, 'Sensor placements'] = best_models['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    best_wrist_models.loc[:, 'Sensor placements'] = best_wrist_models['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    result = pd.concat([best_models, best_wrist_models, result])
    table3_wb = format_for_excel(result)
    result.to_csv(
        output_filepath, float_format='%.2f', index=False)
    table3_wb.save(output_filepath_excel)
    return result


def table_4(input_folder, debug=False):
    output_folder, metrics_file, _, _ = prepare_paths(input_folder, debug=debug)
    output_filepath = os.path.join(output_folder,'figures_and_tables', 'table4.csv')
    output_filepath_excel = os.path.join(output_folder,'figures_and_tables', 'table4.xlsx')
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    summary = pd.read_csv(metrics_file)
    filter_condition = (summary['FEATURE_TYPE'] == 'MO') & (
        summary['NUM_OF_SENSORS'] <= 3)
    table4_data = summary.loc[filter_condition, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'ACTIVITY_AVERAGE',  'ACTIVITY_GROUP_AVERAGE', 'ACTIVITY_IN_GROUP_AVERAGE']]

    table4_data = table4_data.sort_values(by=['ACTIVITY_AVERAGE'], ascending=False).drop_duplicates()
    # best models
    best_models = table4_data.groupby('NUM_OF_SENSORS').apply(lambda rows: rows.head(1)).reset_index(drop=True).sort_values(by=['ACTIVITY_AVERAGE'], ascending=False)
    best_models = best_models.loc[:, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'ACTIVITY_AVERAGE',  'ACTIVITY_GROUP_AVERAGE', 'ACTIVITY_IN_GROUP_AVERAGE']].round(2)
    
    # best models using wrist sensors
    table4_data['USE_DW'] = table4_data['SENSOR_PLACEMENT'].transform(
        lambda s: 'DW' in s.split('_'))
    table4_data['USE_NDW'] = table4_data['SENSOR_PLACEMENT'].transform(
        lambda s: 'NDW' in s.split('_'))
    table4_data['DW_NDW_NONE'] = 'DW'
    table4_data.loc[table4_data['USE_NDW'], 'DW_NDW_NONE'] = 'NDW'
    table4_data['USE_W'] = table4_data['USE_DW'] | table4_data['USE_NDW']
    table4_data.loc[~table4_data['USE_W'], 'DW_NDW_NONE'] = ''
    table4_data.loc[table4_data['USE_DW'] & table4_data['USE_NDW'],'DW_NDW_NONE'] = 'Both W'
    best_wrist_models = table4_data.loc[table4_data['USE_W'] == True,:].groupby('NUM_OF_SENSORS').apply(lambda rows: rows.nlargest(1, columns='ACTIVITY_AVERAGE')).reset_index(drop=True).sort_values(by=['ACTIVITY_AVERAGE'], ascending=False)
    best_wrist_models = best_wrist_models.loc[:, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'ACTIVITY_AVERAGE',  'ACTIVITY_GROUP_AVERAGE', 'ACTIVITY_IN_GROUP_AVERAGE']].round(2)
    # categorized performances
    table4_data['CATEGORY'] = 'W'
    
    # condition 1: A, T, H
    c1 = table4_data['SENSOR_PLACEMENT'].str.contains('A') & table4_data['SENSOR_PLACEMENT'].str.contains('H') & table4_data['SENSOR_PLACEMENT'].str.contains('T')
    table4_data.loc[c1, 'CATEGORY'] = 'A, H, T'
    # condition 2: A, T
    c2 = table4_data['SENSOR_PLACEMENT'].str.contains('A') & table4_data['SENSOR_PLACEMENT'].str.contains('T') & (~table4_data['SENSOR_PLACEMENT'].str.contains('H'))
    table4_data.loc[c2, 'CATEGORY'] = 'A, T'
    # condition 4: A, H
    c4 = table4_data['SENSOR_PLACEMENT'].str.contains('A') & table4_data['SENSOR_PLACEMENT'].str.contains('H') & (~table4_data['SENSOR_PLACEMENT'].str.contains('T'))
    table4_data.loc[c4, 'CATEGORY'] = 'A, H'

    # condition 4: H, T
    c4 = table4_data['SENSOR_PLACEMENT'].str.contains('T') & table4_data['SENSOR_PLACEMENT'].str.contains('H') & (~table4_data['SENSOR_PLACEMENT'].str.contains('A'))
    table4_data.loc[c4, 'CATEGORY'] = 'H, T'

    # condition 5: A
    c5 = table4_data['SENSOR_PLACEMENT'].str.contains('A') & (~table4_data['SENSOR_PLACEMENT'].str.contains('H')) & (~table4_data['SENSOR_PLACEMENT'].str.contains('T'))
    table4_data.loc[c5, 'CATEGORY'] = 'A'

    # condition 6: H
    c6 = table4_data['SENSOR_PLACEMENT'].str.contains('H') & (~table4_data['SENSOR_PLACEMENT'].str.contains('T')) & (~table4_data['SENSOR_PLACEMENT'].str.contains('A'))
    table4_data.loc[c6, 'CATEGORY'] = 'H'

    # condition 7: T
    c7 = table4_data['SENSOR_PLACEMENT'].str.contains('T') & (~table4_data['SENSOR_PLACEMENT'].str.contains('H')) & (~table4_data['SENSOR_PLACEMENT'].str.contains('A'))
    table4_data.loc[c7, 'CATEGORY'] = 'T'

    # condition 8: only W
    c8 = table4_data['SENSOR_PLACEMENT'].str.contains('W') & (~table4_data['SENSOR_PLACEMENT'].str.contains('H')) & (~table4_data['SENSOR_PLACEMENT'].str.contains('A')) & (~table4_data['SENSOR_PLACEMENT'].str.contains('T'))
    table4_data.loc[c8, 'CATEGORY'] = 'W only'

    # categorized for wrist models
    w_categories = table4_data.loc[table4_data['USE_W'], :].groupby(['NUM_OF_SENSORS', 'CATEGORY', 'DW_NDW_NONE']).apply(basic_stat, columns=['ACTIVITY_AVERAGE',  'ACTIVITY_GROUP_AVERAGE', 'ACTIVITY_IN_GROUP_AVERAGE'], method='mean_std').reset_index(drop=False).sort_values(['SORT', 'CATEGORY', 'NUM_OF_SENSORS'], ascending=False).drop(columns=['SORT'])
    
    w_categories['CATEGORY'] = w_categories['DW_NDW_NONE'] + ', ' + w_categories['CATEGORY']
    w_categories = w_categories.drop(columns=['DW_NDW_NONE'])

    # categorize for non-wrist models
    nw_categories = table4_data.loc[~table4_data['USE_W'], :].groupby(['NUM_OF_SENSORS']).apply(basic_stat, columns=['ACTIVITY_AVERAGE',  'ACTIVITY_GROUP_AVERAGE', 'ACTIVITY_IN_GROUP_AVERAGE'], method='mean_std').reset_index(drop=False).sort_values(['SORT', 'NUM_OF_SENSORS'], ascending=False).drop(columns=['SORT'])
    nw_categories.insert(1, 'CATEGORY', 'A, H, T')

    best_models.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Between activity groups', 'Within activity groups']
    best_wrist_models.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Between activity groups', 'Within activity groups']
    w_categories.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Between activity groups', 'Within activity groups']
    nw_categories.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Between activity groups', 'Within activity groups']

    best_models.loc[:, 'Sensor placements'] = best_models['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    best_wrist_models.loc[:, 'Sensor placements'] = best_wrist_models['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    result = pd.concat([best_models, best_wrist_models, w_categories, nw_categories])
    table4_wb = format_for_excel(result)
    result.to_csv(
        output_filepath, float_format='%.2f', index=False)
    table4_wb.save(output_filepath_excel)
    return result


def supplementary_table_1(input_folder, debug=False):
    output_folder, metrics_file, _, _ = prepare_paths(input_folder, debug=debug)
    output_filepath = os.path.join(output_folder,'figures_and_tables', 'supplementary_table1.csv')
    output_filepath_excel = os.path.join(output_folder,'figures_and_tables', 'supplementary_table1.xlsx')
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    summary = pd.read_csv(metrics_file)
    filter_condition = (summary['FEATURE_TYPE'] == 'MO') & (
        summary['NUM_OF_SENSORS'] <= 3)
    table5_data = summary.loc[filter_condition, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'POSTURE_AVERAGE', 'LYING_POSTURE', 'SITTING_POSTURE', 'UPRIGHT_POSTURE']]
    filtered_table5_data = table5_data
    filtered_table5_data.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Lying', 'Sitting', 'Upright']
    filtered_table5_data = filtered_table5_data.sort_values(
        by=['Average'], ascending=False)
    filtered_table5_data.loc[:, 'Sensor placements'] = filtered_table5_data['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    filtered_table5_data = filtered_table5_data.drop_duplicates()
    table5_wb = format_for_excel(filtered_table5_data)
    filtered_table5_data.to_csv(
        output_filepath, float_format='%.2f', index=False)
    table5_wb.save(output_filepath_excel)
    return filtered_table5_data


def supplementary_table_2(input_folder, debug=False):
    output_folder, metrics_file, _, _ = prepare_paths(input_folder, debug=debug)
    output_filepath = os.path.join(output_folder,'figures_and_tables', 'supplementary_table2.csv')
    output_filepath_excel = os.path.join(output_folder,'figures_and_tables', 'supplementary_table2.xlsx')
    os.makedirs(os.path.dirname(output_filepath), exist_ok=True)
    summary = pd.read_csv(metrics_file)
    filter_condition = (summary['FEATURE_TYPE'] == 'MO') & (
        summary['NUM_OF_SENSORS'] <= 3)
    table6_data = summary.loc[filter_condition, [
        'NUM_OF_SENSORS', 'SENSOR_PLACEMENT', 'ACTIVITY_AVERAGE',  'ACTIVITY_GROUP_AVERAGE', 'ACTIVITY_IN_GROUP_AVERAGE']]
    # filtered_table6_data = table6_data.groupby('NUM_OF_SENSORS').apply(
    #     top_and_bottom_n, column='ACTIVITY_AVERAGE', n=5).reset_index(drop=True)
    filtered_table6_data = table6_data
    filtered_table6_data.columns = ['# of sensors', 'Sensor placements',
                                    'Average', 'Between activity groups', 'Within activity groups']
    filtered_table6_data = filtered_table6_data.sort_values(
        by=['Average'], ascending=False)
    filtered_table6_data.loc[:, 'Sensor placements'] = filtered_table6_data['Sensor placements'].transform(
        lambda s: s.replace('_', ', '))
    filtered_table6_data = filtered_table6_data.drop_duplicates()
    table6_wb = format_for_excel(filtered_table6_data)
    filtered_table6_data.to_csv(
        output_filepath, float_format='%.2f', index=False)
    table6_wb.save(output_filepath_excel)
    return filtered_table6_data


def figure_1(input_folder, debug=False):
    rcParams.update(rcParamsDefault)
    output_folder, metrics_file, _, _ = prepare_paths(input_folder, debug=debug)
    rcParams['font.family'] = 'serif'
    rcParams['font.size'] = 12
    rcParams['font.serif'] = ['Times New Roman']
    # setup configurations
    figure_file_extensions = ['.png', '.svg', '.pdf', '.eps']
    output_filepaths = [os.path.join(output_folder, 'figures_and_tables', 'figure1' + extension) for extension in figure_file_extensions]
    os.makedirs(os.path.dirname(output_filepaths[0]), exist_ok=True)

    # read data
    summary = pd.read_csv(metrics_file)

    #  prepare line plot data
    line_plot_data = summary[[
        'NUM_OF_SENSORS', 'FEATURE_TYPE', 'POSTURE_AVERAGE', 'ACTIVITY_AVERAGE']]
    line_plot_data.columns = [
        'Number of sensors', 'Feature set', 'Posture', 'PA']
    line_plot_data.loc[:, 'Feature set'] = line_plot_data['Feature set'].map({
        'M': 'Motion features only',
        'O': 'Orientation related features only',
        'MO': 'Motion + orientation related features'
    })

    line_plot_table = line_plot_data.groupby(['Number of sensors', 'Feature set']).apply(lambda rows: rows[['Posture', 'PA']].mean()).reset_index(drop=False)

    line_plot_table_std = line_plot_data.groupby(['Number of sensors', 'Feature set']).apply(lambda rows: rows[['Posture', 'PA']].std()).reset_index(drop=True)

    line_plot_table_std.columns = ['Posture std', 'PA std']
    line_plot_table = pd.concat((line_plot_table, line_plot_table_std), axis=1)

    # prepare point plot data
    point_plot_data = summary.loc[summary['FEATURE_TYPE'] == 'MO', [
        'SENSOR_PLACEMENT', 'NUM_OF_SENSORS', 'POSTURE_AVERAGE', 'ACTIVITY_AVERAGE']]
    point_plot_data.columns = ['Sensor placements',
                               'Number of sensors', 'Posture', 'PA']

    point_plot_data = pd.melt(point_plot_data, id_vars=['Sensor placements', 'Number of sensors'], value_vars=[
                              'Posture', 'PA'], var_name='Classification task', value_name='F1-score')
    point_plot_data['Include dominant wrist'] = point_plot_data['Sensor placements'].transform(
        lambda s: 'DW' in s.split('_'))
    point_plot_data['Include nondominant wrist'] = point_plot_data['Sensor placements'].transform(
        lambda s: 'NDW' in s.split('_'))
    point_plot_data['Include wrists'] = point_plot_data['Include dominant wrist'] | point_plot_data['Include nondominant wrist']

    point_plot_data = point_plot_data.sort_values(
        by=['Number of sensors', 'Classification task', 'F1-score'], ascending=True)

    # point_plot_data = point_plot_data.groupby(['Number of sensors', 'Classification task']).apply(
    #     lambda rows: pd.concat((rows.head(5), rows.tail(5))))
    point_plot_data = point_plot_data.reset_index(drop=True).drop_duplicates()

    # draw plots
    g, axes = plt.subplots(2, 2, figsize=(8, 8))
    sns.set_context("paper")
    sns.set_style("white")
    for task, index in zip(['Posture', 'PA'], [0, 1]):
        sns.set(rc={"lines.linewidth": 0.8,
                    "font.family": ['serif'],
                    "font.serif": ['Times New Roman'],
                    "font.size": 12
                    })
        # draw swarm and line for MO feature set
        axes[index][0].set_ylim(0, 1.2)
        axes[index][0].set_xlim(0, 7)
        axes[index][0].yaxis.grid(linestyle='--')
        swarm_data = point_plot_data.loc[point_plot_data['Classification task'] == task, :]
        if task == 'Posture':
            swarm_data_wrist = swarm_data.loc[swarm_data['Include wrists'] == True, :]
            swarm_data_nonwrist = swarm_data.loc[swarm_data['Include wrists'] == False, :]
        else:
            swarm_data_wrist = swarm_data.loc[swarm_data['Include dominant wrist'] == True, :]
            swarm_data_nonwrist = swarm_data.loc[swarm_data['Include dominant wrist'] == False, :]
        line_data_mo = line_plot_data.loc[line_plot_data['Feature set'] == 'Motion + orientation related features', [
            'Number of sensors', 'Feature set', task]].rename(columns={task: 'F1-score'})
        for x in [0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.5]:
            axes[index][0].axvline(x=x, color='0.75')
        sns.swarmplot(x='Number of sensors', y='F1-score', data=swarm_data,
                      ax=axes[index][0], linewidth=1, hue='Include wrists', palette=sns.color_palette('Greys', n_colors=2), size=4)
        sns.pointplot(x='Number of sensors', y='F1-score',
                      data=line_data_mo, ax=axes[index][0], color='gray', marker='x', capsize=0.1, errwidth=0, hue='Feature set')
        legend_handles = axes[index][0].legend_.legendHandles
        legend_handles[2] = axes[index][0].lines[7]
        if task == 'Posture':
            axes[index][0].legend(handles=legend_handles, labels=["Models without wrist sensors", "Models with wrist sensors", "Motion + orientation features"],
                                  frameon=True, loc='lower right', framealpha=1, fancybox=False, facecolor='white', edgecolor='black', shadow=None)
        else:
            axes[index][0].legend(handles=legend_handles, labels=["Models without wrist sensors", "Models with wrist sensors", "Motion + orientation features"],
                                  frameon=True, loc='lower right', framealpha=1, fancybox=False, facecolor='white', edgecolor='black', shadow=None)

        # draw line for other feature set

        line_data_others = line_plot_data.loc[:, [
            'Number of sensors', 'Feature set', task]].rename(columns={task: 'F1-score'})
        sns.pointplot(x='Number of sensors', y='F1-score', data=line_data_others,
                      dodge=True, ax=axes[index][1], hue='Feature set', hue_order=['Motion + orientation related features', 'Motion features only', 'Orientation related features only'], palette='Greys_r', linestyles=['-', '--', '-.'], markers='x', errwidth=0)
        axes[index][1].legend(handles=[axes[index][1].lines[0], axes[index][1].lines[8], axes[index][1].lines[16]], labels=["Motion + orientation features", "Motion features", "Orientation features"],
                              frameon=True, loc='lower right', framealpha=1, fancybox=False, facecolor='white', edgecolor='black', shadow=None)
        axes[index][1].set_ylim(0, 1.2)
        axes[index][1].set_yticklabels([])
        axes[index][1].yaxis.set_major_formatter(plt.NullFormatter())
        axes[index][1].set_ylabel('')
        axes[index][1].yaxis.grid(linestyle='--')
        # axes[index][1].spines['left'].set_color('grey')

    g.subplots_adjust(wspace=0.04, hspace=0.35)
    plt.figtext(0.5, 0.49, '(a) Posture recognition performance',
                ha='center', va='top')
    plt.figtext(0.5, 0.05, '(b) PA recognition performance',
                ha='center', va='top')
    # plt.show()
    # save figure in different formats
    for output_filepath in output_filepaths:
        print('save ' + output_filepath)
        plt.savefig(output_filepath, dpi=300, orientation='landscape')

    # save associated table
    output_filepath = os.path.join(os.path.dirname(output_filepaths[0]), 'figure1.csv')
    line_plot_table.to_csv(output_filepath, index=False, float_format='%0.3f')
    plt.close(fig=g)
    return g

def figure_2(input_folder, debug=False):
    rcParams.update(rcParamsDefault)
    output_folder, metrics_file, prediction_set_file, confusion_matrix_file = prepare_paths(input_folder, debug=debug)
    # prepare confusion matrix
    abbr_labels = get_pa_abbr_labels(input_folder)
    
    conf_df = pd.read_csv(confusion_matrix_file)
    conf_df = conf_df.rename(columns={conf_df.columns[0]: 'Ground Truth'})
    conf_df = conf_df.set_index(conf_df.columns[0])
    conf_df.columns = abbr_labels
    conf_df.index = abbr_labels
    # plot confusion matrix
    rcParams['font.family'] = 'serif'
    rcParams['font.size'] = 10
    rcParams['font.serif'] = ['Times New Roman']
    fig, ax = plt.subplots(figsize=(8, 8))
    sns.set_style({
        'font.family': 'serif',
        'font.size': 10
    })
    g = sns.heatmap(conf_df, annot=True, cmap="Greys",
                    cbar=False, fmt='d', robust=True, linewidths=0.2)
    g.set(xlabel="Prediction", ylabel="Ground truth")
    plt.tight_layout()

    # save plot
    figure_file_extensions = ['.png', '.svg', '.pdf', '.eps']
    output_filepaths = [os.path.join(output_folder,'figures_and_tables', 'figure2' + extension) for extension in figure_file_extensions]
    os.makedirs(os.path.dirname(output_filepaths[0]), exist_ok=True)
    for output_filepath in output_filepaths:
        plt.savefig(output_filepath, dpi=300, orientation='landscape')

    # prepare plot associated table
    labels = get_pa_labels(input_folder)
    prediction_set = pd.read_csv(prediction_set_file, parse_dates=[
                                 0, 1], infer_datetime_format=True)
    
    f1_scores_per_activity = f1_score(prediction_set['ACTIVITY'], prediction_set['ACTIVITY_PREDICTION'], labels=labels, average=None)
    f1_df = pd.DataFrame(data=f1_scores_per_activity, index=labels, columns = ['F1_score'])
    f1_df = f1_df.sort_values(['F1_score'])
    f1_df = f1_df.loc[f1_df['F1_score']<=0.4,:]
    f1_df.index = [abbr_labels[labels.index(i)] for i in f1_df.index]
    cases = []
    for l in f1_df.index:
        cases.append(top_n_misclassified_classes(conf_df, l, n=3))
    mis_cases = pd.concat(cases, axis=1).transpose()
    result = pd.concat((f1_df.reset_index(drop=False), mis_cases), axis=1)
    result.columns = ['Activity', 'F1 score', 'Misclassifications', '', '']

    # save table
    output_filepath = os.path.join(output_folder,'figures_and_tables', 'figure2.csv')
    result.to_csv(output_filepath, float_format='%.2f', index=False)
    plt.close(fig=fig)
    return fig


def dataset_summary(input_folder, debug=False):
    output_folder, _, _, _ = prepare_paths(input_folder, debug=debug)
    exception_file = os.path.join(input_folder, "MetaCrossParticipants", 'pid_exceptions.csv')
    offset_mapping_file = os.path.join(input_folder, "MetaCrossParticipants", 'offset_mapping.csv')
    orientation_correction_file = os.path.join(input_folder, 'MetaCrossParticipants', 'orientation_corrections.csv')
    subject_file = os.path.join(input_folder, "MetaCrossParticipants", 'subjects.csv')
    class_file = os.path.join(output_folder, 'muss.class.csv')
    subjects = pd.read_csv(subject_file, header=0)
    exceptions = pd.read_csv(exception_file, header=0)
    offset_mapping = pd.read_csv(offset_mapping_file, header=0)
    classes = pd.read_csv(class_file, header=0, parse_dates=[0,1], infer_datetime_format=True)
    orientation_corrections = pd.read_csv(orientation_correction_file, header=0)
    selection = ~subjects.PID.isin(exceptions.PID)
    selected_subjects = subjects.loc[selection,:]
    selection = ~offset_mapping.PID.isin(exceptions.PID)
    selected_offset_mapping = offset_mapping.loc[selection, :]
    selection = ~orientation_corrections.PID.isin(exceptions.PID)
    selected_orientation_corrections = orientation_corrections.loc[selection,:]
    summary = dict()
    summary['mean_age'] = selected_subjects.AGE.mean()
    summary['std_age'] = selected_subjects.AGE.std()
    summary['male'] = np.sum(selected_subjects.GENDER == 'M')
    summary['female'] = np.sum(selected_subjects.GENDER == 'F')
    summary['mean_bmi'] = selected_subjects.iloc[:, 5].mean()
    summary['std_bmi'] = selected_subjects.iloc[:, 5].std()
    summary['offset_percentage'] = np.sum(selected_offset_mapping.iloc[:, 1] != 0) / 42.0
    summary['mean_offset'] = selected_offset_mapping.iloc[:, 1].mean()
    summary['std_offset'] = selected_offset_mapping.iloc[:, 1].std()
    summary['misplace_percentage'] = selected_orientation_corrections.shape[0] / (42.0 * 7)
    # summary['misplace_ankle_percentage'] = np.sum(selected_orientation_corrections.SENSOR_PLACEMENT.str.contains('ankle|waist|thigh')) / selected_orientation_corrections.shape[0]
    summary['total_samples'] = np.sum(~classes.ACTIVITY.isin(['Transition', 'Unknown']))
    summary['total_activities'] = len(classes.ACTIVITY.unique()) - 2
    summary['upright_samples'] = np.sum(classes.POSTURE == 'Upright')
    summary['sitting_samples'] = np.sum(classes.POSTURE == 'Sitting')
    summary['lying_samples'] = np.sum(classes.POSTURE == 'Lying')
    activity_samples = classes.ACTIVITY.value_counts()
    activity_samples = activity_samples.to_frame()
    activity_samples['DURATION(MIN)'] = activity_samples.values * 12.8 / 60.0
    ag_samples = classes.ACTIVITY_GROUP.value_counts()
    ag_samples = ag_samples.to_frame()
    ag_samples['DURATION(MIN)'] = ag_samples.values * 12.8 / 60.0
    summary_output = os.path.join(output_folder, 'figures_and_tables', 'dataset_stats.csv')
    summary_df = pd.DataFrame.from_dict(summary, orient='index')
    summary_df.to_csv(summary_output, float_format='%.3f', index=True, header=False)
    activity_samples_output = os.path.join(output_folder, 'figures_and_tables', 'activity_samples_stats.csv')
    activity_samples.to_csv(activity_samples_output, float_format='%.3f', index=True, header=True)
    ag_samples_output = os.path.join(output_folder, 'figures_and_tables', 'activity_group_samples_stats.csv')
    ag_samples.to_csv(ag_samples_output, float_format='%.3f', index=True, header=True)
    return summary_df, activity_samples, ag_samples

def numbers_in_abstract(input_folder, debug=False):
    output_folder, metrics_file, _, _ = prepare_paths(input_folder, debug=debug)
    summary = pd.read_csv(metrics_file)
    filter_condition = (summary.NUM_OF_SENSORS == 2) & (~summary.SENSOR_PLACEMENT.str.contains('W')) & (summary.FEATURE_TYPE == 'MO')
    two_non_wrist_models = summary.loc[filter_condition,:]
    print(two_non_wrist_models.SENSOR_PLACEMENT)

    filter_condition2 = (summary.NUM_OF_SENSORS == 2) & (summary.SENSOR_PLACEMENT.str.contains('W')) & (summary.FEATURE_TYPE == 'MO')
    two_wrist_models = summary.loc[filter_condition2,:]
    print(two_wrist_models.SENSOR_PLACEMENT)

    filter_condition3 = two_wrist_models.SENSOR_PLACEMENT != 'DW_NDW'
    two_wrist_and_non_wrist_models = two_wrist_models.loc[filter_condition3,:]
    print(two_wrist_and_non_wrist_models.SENSOR_PLACEMENT)


    filter_condition4 = (summary.NUM_OF_SENSORS == 2) & (~summary.SENSOR_PLACEMENT.isin(two_wrist_and_non_wrist_models.SENSOR_PLACEMENT)) & (summary.FEATURE_TYPE == 'MO')
    other_two_models = summary.loc[filter_condition4,:]
    print(other_two_models.SENSOR_PLACEMENT)

    result = dict()
    result['TWO_NON_WRIST'] = {
        'POSTURE_MEAN': two_non_wrist_models.POSTURE_AVERAGE.mean(),
        'POSTURE_STD': two_non_wrist_models.POSTURE_AVERAGE.std(),
        'PA_MEAN': two_non_wrist_models.ACTIVITY_AVERAGE.mean(),
        'PA_STD': two_non_wrist_models.ACTIVITY_AVERAGE.std()
    }
    result['TWO_WRIST_AND_NON_WRIST'] = {
        'POSTURE_MEAN': two_wrist_and_non_wrist_models.POSTURE_AVERAGE.mean(),
        'POSTURE_STD': two_wrist_and_non_wrist_models.POSTURE_AVERAGE.std(),
        'PA_MEAN': two_wrist_and_non_wrist_models.ACTIVITY_AVERAGE.mean(),
        'PA_STD': two_wrist_and_non_wrist_models.ACTIVITY_AVERAGE.std()
    }
    result['TWO_WRIST'] = {
        'POSTURE_MEAN': two_wrist_models.POSTURE_AVERAGE.mean(),
        'POSTURE_STD': two_wrist_models.POSTURE_AVERAGE.std(),
        'PA_MEAN': two_wrist_models.ACTIVITY_AVERAGE.mean(),
        'PA_STD': two_wrist_models.ACTIVITY_AVERAGE.std()
    }
    result['OTHER_THAN_TWO_WRIST_AND_NON_WRIST'] = {
        'POSTURE_MEAN': other_two_models.POSTURE_AVERAGE.mean(),
        'POSTURE_STD': other_two_models.POSTURE_AVERAGE.std(),
        'PA_MEAN': other_two_models.ACTIVITY_AVERAGE.mean(),
        'PA_STD': other_two_models.ACTIVITY_AVERAGE.std()
    }
    result = pd.DataFrame(result).transpose()
    output_filepath = os.path.join(output_folder,'figures_and_tables', 'numbers_in_abstract.csv')
    result.to_csv(output_filepath, index=True, float_format='%.3f')
    return result

def prepare_paths(input_folder, output_folder=None, debug=False):
    if output_folder is None:
        output_folder = generate_run_folder(input_folder, debug=debug)
    os.makedirs(output_folder, exist_ok=True)
    metrics_file = os.path.join(output_folder, 'muss.metrics.csv')
    figure_2_predictions = os.path.join(
        output_folder, 'predictions', 'DW_DT.MO.prediction.csv')
    figure_2_confusion_matrix = os.path.join(
        output_folder, 'confusion_matrices', 'DW_DT.MO.confusion_matrix.csv')
    return output_folder, metrics_file, figure_2_predictions, figure_2_confusion_matrix

def main(input_folder, *, output_folder=None, debug=False, force=True):
    """Generate figures and tables used in the paper

    :param input_folder: Folder path of input raw dataset
    :param output_folder: Auto path if None
    :param debug: Use this flag to output results to 'debug_run' folder
    """
    figure_folder = os.path.join(output_folder,'figures_and_tables')
    if not force and os.path.exists(figure_folder):
        logging.info('Figures exist, skip regenerating them...')
        return figure_folder
    table_3(input_folder, output_folder=output_folder, debug=debug)
    table_4(input_folder, output_folder=output_folder, debug=debug)
    figure_1(input_folder, output_folder=output_folder, debug=debug)
    figure_2(input_folder, output_folder=output_folder, debug=debug)
    supplementary_table_1(input_folder, output_folder=output_folder, debug=debug)
    supplementary_table_2(input_folder, output_folder=output_folder, debug=debug)
    dataset_summary(input_folder, output_folder=output_folder, debug=debug)
    numbers_in_abstract(input_folder, output_folder=output_folder, debug=debug)
    output_folder = generate_run_folder(input_folder, output_folder=output_folder, debug=debug)
    return figure_folder

if __name__ == '__main__':
    run(main)
    
    
    